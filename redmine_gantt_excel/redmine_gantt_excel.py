#!/usr/bin/env python3
"""
Redmine issues -> Gantt-style Excel generator

Requirements:
  pip install requests openpyxl

Example:
  python redmine_gantt_excel.py \
      --config redmine_gantt.ini \
      --project-id 123 \
      --output redmine_gantt.xlsx

INI example:
  [redmine]
  base_url = https://redmine.example.com
  api_key = YOUR_API_KEY

Notes:
- This script reads issues from Redmine REST API and generates:
  1) A Gantt-style worksheet
  2) A raw issue worksheet
- It uses issue start_date / due_date to render bars.
- Parent-child hierarchy is shown via indent and WBS numbering.
- Relations are listed as text in a separate column.
- CLI arguments override INI values.
"""

from __future__ import annotations

import argparse
import configparser
import datetime as dt
import math
import sys
import time
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import urljoin

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


DATE_FMT = "%Y-%m-%d"
DEFAULT_CONFIG_PATH = "redmine_gantt.ini"


@dataclass
class Issue:
    issue_id: int
    project_id: Optional[int]
    project_name: str
    tracker_name: str
    status_name: str
    priority_name: str
    subject: str
    assigned_to: str
    fixed_version: str
    parent_id: Optional[int]
    start_date: Optional[dt.date]
    due_date: Optional[dt.date]
    done_ratio: int
    estimated_hours: Optional[float]
    project_path: str = ""
    relations: List[Dict[str, Any]] = field(default_factory=list)
    custom_fields: Dict[str, Any] = field(default_factory=dict)
    raw: Dict[str, Any] = field(default_factory=dict)
    depth: int = 0
    wbs: str = ""


class RedmineClient:
    def __init__(self, base_url: str, api_key: str, timeout: int = 30) -> None:
        self.base_url = base_url.rstrip("/") + "/"
        self.session = requests.Session()
        self.session.headers.update({
            "X-Redmine-API-Key": api_key,
            "Accept": "application/json",
            "User-Agent": "redmine-gantt-excel/1.1",
        })
        self.timeout = timeout

    def _get(self, path: str, params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        url = urljoin(self.base_url, path.lstrip("/"))
        resp = self.session.get(url, params=params or {}, timeout=self.timeout)
        try:
            resp.raise_for_status()
        except requests.HTTPError as e:
            msg = f"HTTP error for {resp.url}: {resp.status_code} {resp.text[:400]}"
            raise RuntimeError(msg) from e
        try:
            return resp.json()
        except ValueError as e:
            raise RuntimeError(f"Invalid JSON response from {resp.url}") from e

    def fetch_issues(
        self,
        project_id: Optional[int],
        status_id: str,
        limit: int,
        trackers: Optional[str],
        assigned_to_id: Optional[str],
        version_id: Optional[str],
        subproject_mode: Optional[str],
        fallback_relations: bool,
        sleep_sec: float,
    ) -> List[Issue]:
        issues: List[Issue] = []
        offset = 0
        total_count = None

        while True:
            params: Dict[str, Any] = {
                "offset": offset,
                "limit": limit,
                "status_id": status_id,
                "sort": "start_date:asc,due_date:asc,id:asc",
                "include": "relations",
            }
            if project_id is not None:
                params["project_id"] = project_id
            if trackers:
                params["tracker_id"] = trackers
            if assigned_to_id:
                params["assigned_to_id"] = assigned_to_id
            if version_id:
                params["fixed_version_id"] = version_id
            if subproject_mode:
                params["subproject_id"] = subproject_mode

            payload = self._get("issues.json", params=params)
            batch = payload.get("issues", [])
            total_count = payload.get("total_count", total_count)

            for item in batch:
                issue = self._parse_issue(item)
                if fallback_relations and not issue.relations:
                    issue.relations = self.fetch_issue_relations(issue.issue_id, sleep_sec=sleep_sec)
                issues.append(issue)

            offset += len(batch)
            if not batch:
                break
            if total_count is not None and offset >= total_count:
                break
            time.sleep(sleep_sec)

        return issues

    def fetch_issue_relations(self, issue_id: int, sleep_sec: float = 0.0) -> List[Dict[str, Any]]:
        payload = self._get(f"issues/{issue_id}.json", params={"include": "relations"})
        issue = payload.get("issue", {})
        time.sleep(sleep_sec)
        return issue.get("relations", []) or []

    def fetch_projects(self, limit: int = 100, sleep_sec: float = 0.0) -> List[Dict[str, Any]]:
        projects: List[Dict[str, Any]] = []
        offset = 0
        total_count = None

        while True:
            payload = self._get("projects.json", params={"offset": offset, "limit": limit})
            batch = payload.get("projects", [])
            total_count = payload.get("total_count", total_count)
            projects.extend(batch)

            offset += len(batch)
            if not batch:
                break
            if total_count is not None and offset >= total_count:
                break
            time.sleep(sleep_sec)

        return projects

    @staticmethod
    def _parse_issue(item: Dict[str, Any]) -> Issue:
        def parse_date(text: Optional[str]) -> Optional[dt.date]:
            if not text:
                return None
            return dt.datetime.strptime(text, DATE_FMT).date()

        def name_of(obj: Optional[Dict[str, Any]]) -> str:
            if not obj:
                return ""
            return str(obj.get("name", ""))

        custom_fields = {
            str(cf.get("name", "")): cf.get("value")
            for cf in item.get("custom_fields", [])
            if cf.get("name")
        }

        project = item.get("project") or {}

        return Issue(
            issue_id=int(item["id"]),
            project_id=project.get("id"),
            project_name=name_of(project),
            tracker_name=name_of(item.get("tracker")),
            status_name=name_of(item.get("status")),
            priority_name=name_of(item.get("priority")),
            subject=str(item.get("subject", "")),
            assigned_to=name_of(item.get("assigned_to")),
            fixed_version=name_of(item.get("fixed_version")),
            parent_id=(item.get("parent") or {}).get("id"),
            start_date=parse_date(item.get("start_date")),
            due_date=parse_date(item.get("due_date")),
            done_ratio=int(item.get("done_ratio") or 0),
            estimated_hours=(float(item["estimated_hours"]) if item.get("estimated_hours") is not None else None),
            relations=item.get("relations", []) or [],
            custom_fields=custom_fields,
            raw=item,
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a Gantt-style Excel workbook from Redmine issues.")
    parser.add_argument("--config", default=DEFAULT_CONFIG_PATH, help=f"INI config path (default: {DEFAULT_CONFIG_PATH})")
    parser.add_argument("--base-url", default=None, help="Redmine base URL, e.g. https://redmine.example.com")
    parser.add_argument("--api-key", default=None, help="Redmine API key")
    parser.add_argument("--project-id", type=int, default=None, help="Project numeric ID")
    parser.add_argument("--status-id", default="*", help="Redmine status filter. Use '*' for all.")
    parser.add_argument("--tracker-id", default=None, help="Comma-separated tracker IDs")
    parser.add_argument("--assigned-to-id", default=None, help="Assigned user id or 'me'")
    parser.add_argument("--fixed-version-id", default=None, help="Fixed version id")
    parser.add_argument(
        "--subproject-id",
        default=None,
        help="Subproject filter. Example: '!*' means only the project itself and no subprojects.",
    )
    parser.add_argument("--limit", type=int, default=100, help="API page size")
    parser.add_argument("--output", default="redmine_gantt.xlsx", help="Output xlsx path")
    parser.add_argument("--sheet-name", default="Gantt", help="Main sheet name")
    parser.add_argument("--from-date", default=None, help="Override chart start date (YYYY-MM-DD)")
    parser.add_argument("--to-date", default=None, help="Override chart end date (YYYY-MM-DD)")
    parser.add_argument(
        "--timeline-mode",
        choices=["day", "week"],
        default="day",
        help="Timeline unit. 'day' is more detailed, 'week' is more compact.",
    )
    parser.add_argument(
        "--fallback-relations",
        action="store_true",
        help="Fetch relations per issue when list API does not include them consistently.",
    )
    parser.add_argument("--sleep-sec", type=float, default=0.0, help="Sleep between API requests")
    parser.add_argument(
        "--include-no-date-issues",
        action="store_true",
        help="Include issues with no start/due date in the table (without bar).",
    )
    return parser.parse_args()


def read_ini_config(config_path: str) -> configparser.ConfigParser:
    config = configparser.ConfigParser()
    path = Path(config_path)
    if path.exists():
        config.read(path, encoding="utf-8")
    return config


def get_config_value(config: configparser.ConfigParser, section: str, key: str) -> Optional[str]:
    if config.has_option(section, key):
        value = config.get(section, key).strip()
        return value or None
    return None


def resolve_connection_settings(args: argparse.Namespace, config: configparser.ConfigParser) -> Tuple[str, str]:
    base_url = args.base_url or get_config_value(config, "redmine", "base_url")
    api_key = args.api_key or get_config_value(config, "redmine", "api_key")

    missing = []
    if not base_url:
        missing.append("base_url")
    if not api_key:
        missing.append("api_key")

    if missing:
        joined = ", ".join(missing)
        raise ValueError(
            f"Missing required setting(s): {joined}. "
            f"Pass them via CLI or set them in [{'redmine'}] section of {args.config}."
        )

    return base_url, api_key


def build_project_paths(projects: List[Dict[str, Any]]) -> Dict[int, str]:
    by_id = {int(p["id"]): p for p in projects if p.get("id") is not None}
    resolved: Dict[int, str] = {}

    def name_of_project(project_id: Optional[int], stack: Optional[set[int]] = None) -> str:
        if project_id is None:
            return ""
        if project_id in resolved:
            return resolved[project_id]
        project = by_id.get(project_id)
        if not project:
            return ""
        if stack is None:
            stack = set()
        if project_id in stack:
            return str(project.get("name", ""))
        stack.add(project_id)

        parent = project.get("parent") or {}
        parent_id = parent.get("id")
        parent_path = name_of_project(parent_id, stack)
        current_name = str(project.get("name", ""))
        path = f"{parent_path} / {current_name}" if parent_path else current_name
        resolved[project_id] = path
        stack.remove(project_id)
        return path

    for project_id in by_id:
        name_of_project(project_id)

    return resolved


def attach_project_paths(issues: List[Issue], project_paths: Dict[int, str]) -> None:
    for issue in issues:
        if issue.project_id is not None:
            issue.project_path = project_paths.get(issue.project_id, issue.project_name)
        else:
            issue.project_path = issue.project_name


def compute_hierarchy(issues: List[Issue]) -> List[Issue]:
    by_id = {i.issue_id: i for i in issues}
    children: Dict[Optional[int], List[Issue]] = defaultdict(list)
    roots: List[Issue] = []

    for issue in issues:
        if issue.parent_id and issue.parent_id in by_id:
            children[issue.parent_id].append(issue)
        else:
            roots.append(issue)

    def sort_key(x: Issue) -> Tuple:
        start = x.start_date or dt.date.max
        due = x.due_date or dt.date.max
        return (start, due, x.issue_id)

    roots.sort(key=sort_key)
    for key in list(children.keys()):
        children[key].sort(key=sort_key)

    ordered: List[Issue] = []

    def visit(node: Issue, prefix: str, depth: int) -> None:
        node.depth = depth
        node.wbs = prefix
        ordered.append(node)
        for idx, ch in enumerate(children.get(node.issue_id, []), start=1):
            visit(ch, f"{prefix}.{idx}", depth + 1)

    for idx, root in enumerate(roots, start=1):
        visit(root, str(idx), 0)

    return ordered


def detect_date_range(issues: List[Issue], include_no_date_issues: bool) -> Tuple[List[Issue], dt.date, dt.date]:
    usable: List[Issue] = []
    min_dates: List[dt.date] = []
    max_dates: List[dt.date] = []

    for issue in issues:
        if issue.start_date and issue.due_date:
            usable.append(issue)
            min_dates.append(issue.start_date)
            max_dates.append(issue.due_date)
        elif include_no_date_issues:
            usable.append(issue)

    if not min_dates or not max_dates:
        today = dt.date.today()
        return usable, today, today + dt.timedelta(days=30)

    start = min(min_dates)
    end = max(max_dates)
    return usable, start, end


def override_date_range(start: dt.date, end: dt.date, from_date: Optional[str], to_date: Optional[str]) -> Tuple[dt.date, dt.date]:
    if from_date:
        start = dt.datetime.strptime(from_date, DATE_FMT).date()
    if to_date:
        end = dt.datetime.strptime(to_date, DATE_FMT).date()
    if start > end:
        raise ValueError("from-date must be <= to-date")
    return start, end


def daterange(start: dt.date, end: dt.date) -> Iterable[dt.date]:
    current = start
    while current <= end:
        yield current
        current += dt.timedelta(days=1)


def week_start(d: dt.date) -> dt.date:
    return d - dt.timedelta(days=d.weekday())


def build_timeline(start: dt.date, end: dt.date, mode: str) -> List[Tuple[dt.date, dt.date]]:
    if mode == "day":
        return [(d, d) for d in daterange(start, end)]

    timeline: List[Tuple[dt.date, dt.date]] = []
    cursor = week_start(start)
    while cursor <= end:
        timeline.append((cursor, cursor + dt.timedelta(days=6)))
        cursor += dt.timedelta(days=7)
    return timeline


def relation_text(issue: Issue) -> str:
    if not issue.relations:
        return ""
    parts = []
    for rel in issue.relations:
        rel_type = str(rel.get("relation_type", "relates"))
        other_id = rel.get("issue_to_id")
        if other_id is None and rel.get("issue_id") != issue.issue_id:
            other_id = rel.get("issue_id")
        delay = rel.get("delay")
        suffix = f"(+{delay}d)" if delay not in (None, "", 0) else ""
        parts.append(f"{rel_type} #{other_id}{suffix}")
    return ", ".join(parts)


def apply_common_styles(ws: Worksheet) -> None:
    thin = Side(border_style="thin", color="D9D9D9")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.row >= 3:
                cell.alignment = Alignment(vertical="center")


def write_gantt_sheet(
    wb: Workbook,
    sheet_name: str,
    issues: List[Issue],
    timeline: List[Tuple[dt.date, dt.date]],
    timeline_mode: str,
) -> Worksheet:
    ws = wb.active
    ws.title = sheet_name

    fixed_cols = [
        ("WBS", 10),
        ("ID", 8),
        ("ProjectPath", 28),
        ("Subject", 52),
        ("Tracker", 14),
        ("Status", 14),
        ("Priority", 12),
        ("Assignee", 18),
        ("Version", 16),
        ("Start", 12),
        ("Due", 12),
        ("Progress", 10),
        ("Est.Hours", 10),
        ("Parent", 8),
        ("Relations", 30),
    ]
    timeline_start_col = len(fixed_cols) + 1

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    month_fill = PatternFill("solid", fgColor="D9EAF7")
    weekend_fill = PatternFill("solid", fgColor="F2F2F2")
    today_fill = PatternFill("solid", fgColor="FFF2CC")
    bar_fill = PatternFill("solid", fgColor="9DC3E6")
    progress_fill = PatternFill("solid", fgColor="5B9BD5")
    summary_bar_fill = PatternFill("solid", fgColor="A9D18E")
    nodate_fill = PatternFill("solid", fgColor="FAFAFA")
    parent_subject_fill = PatternFill("solid", fgColor="E2F0D9")

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    for idx, (name, width) in enumerate(fixed_cols, start=1):
        ws.cell(row=3, column=idx, value=name)
        ws.column_dimensions[get_column_letter(idx)].width = width
        ws.cell(row=3, column=idx).fill = header_fill
        ws.cell(row=3, column=idx).font = header_font
        ws.cell(row=3, column=idx).alignment = Alignment(horizontal="center", vertical="center")

    today = dt.date.today()

    if timeline_mode == "day":
        month_spans: List[Tuple[int, int, str]] = []
        current_label = None
        start_col = timeline_start_col
        for idx, (d1, _) in enumerate(timeline, start=timeline_start_col):
            label = d1.strftime("%Y-%m")
            if label != current_label:
                if current_label is not None:
                    month_spans.append((start_col, idx - 1, current_label))
                current_label = label
                start_col = idx
            cell = ws.cell(row=2, column=idx, value=f"{d1.day}\n{d1.strftime('%a')}")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if d1.weekday() >= 5:
                cell.fill = weekend_fill
            if d1 == today:
                cell.fill = today_fill
            ws.column_dimensions[get_column_letter(idx)].width = 4.3
        if current_label is not None:
            month_spans.append((start_col, timeline_start_col + len(timeline) - 1, current_label))

        for start_c, end_c, label in month_spans:
            ws.merge_cells(start_row=1, start_column=start_c, end_row=1, end_column=end_c)
            cell = ws.cell(row=1, column=start_c, value=label)
            cell.fill = month_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
    else:
        for idx, (d1, d2) in enumerate(timeline, start=timeline_start_col):
            month_label = d1.strftime("%Y-%m")
            week_label = f"W{int(d1.strftime('%W')):02d}\n{d1.strftime('%m/%d')}-{d2.strftime('%m/%d')}"
            ws.cell(row=1, column=idx, value=month_label)
            ws.cell(row=1, column=idx).fill = month_fill
            ws.cell(row=1, column=idx).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=1, column=idx).font = Font(bold=True)
            ws.cell(row=2, column=idx, value=week_label)
            ws.cell(row=2, column=idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if d1 <= today <= d2:
                ws.cell(row=2, column=idx).fill = today_fill
            ws.column_dimensions[get_column_letter(idx)].width = 10

    for col in range(1, timeline_start_col):
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        title = "Redmine Gantt"
        if col == 1:
            ws.cell(row=1, column=col, value=title)
            ws.cell(row=1, column=col).font = Font(bold=True, size=12)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=col).fill = month_fill

    parent_ids = {i.parent_id for i in issues if i.parent_id is not None}

    for row_idx, issue in enumerate(issues, start=4):
        ws.row_dimensions[row_idx].height = 20
        values = [
            issue.wbs,
            issue.issue_id,
            issue.project_path,
            ("    " * issue.depth) + issue.subject,
            issue.tracker_name,
            issue.status_name,
            issue.priority_name,
            issue.assigned_to,
            issue.fixed_version,
            issue.start_date.isoformat() if issue.start_date else "",
            issue.due_date.isoformat() if issue.due_date else "",
            f"{issue.done_ratio}%",
            issue.estimated_hours if issue.estimated_hours is not None else "",
            issue.parent_id or "",
            relation_text(issue),
        ]
        for col_idx, value in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in (3, 4, 15):
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(col_idx in (3, 15)))
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")

        if issue.issue_id in parent_ids:
            ws.cell(row=row_idx, column=4).font = Font(bold=True)
            ws.cell(row=row_idx, column=4).fill = parent_subject_fill

        if not (issue.start_date and issue.due_date):
            for col_idx in range(1, len(fixed_cols) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = nodate_fill
            continue

        if timeline_mode == "day":
            bar_start = issue.start_date
            bar_end = issue.due_date
            total_days = (bar_end - bar_start).days + 1
            progress_days = max(0, math.ceil(total_days * (issue.done_ratio / 100.0)))
            progress_end = bar_start + dt.timedelta(days=max(progress_days - 1, -1)) if progress_days > 0 else None

            for col_offset, (d1, _) in enumerate(timeline, start=timeline_start_col):
                cell = ws.cell(row=row_idx, column=col_offset)
                if d1.weekday() >= 5:
                    cell.fill = weekend_fill
                if issue.start_date <= d1 <= issue.due_date:
                    cell.fill = summary_bar_fill if issue.issue_id in parent_ids else bar_fill
                if progress_end and issue.start_date <= d1 <= progress_end:
                    cell.fill = progress_fill
                if d1 == today and cell.fill.patternType is None:
                    cell.fill = today_fill
        else:
            total_days = (issue.due_date - issue.start_date).days + 1
            progress_days = max(0, math.ceil(total_days * (issue.done_ratio / 100.0)))
            progress_end = issue.start_date + dt.timedelta(days=max(progress_days - 1, -1)) if progress_days > 0 else None
            for col_offset, (w_start, w_end) in enumerate(timeline, start=timeline_start_col):
                cell = ws.cell(row=row_idx, column=col_offset)
                overlaps_bar = not (issue.due_date < w_start or issue.start_date > w_end)
                overlaps_progress = progress_end is not None and not (progress_end < w_start or issue.start_date > w_end)
                if overlaps_bar:
                    cell.fill = summary_bar_fill if issue.issue_id in parent_ids else bar_fill
                if overlaps_progress:
                    cell.fill = progress_fill
                if w_start <= today <= w_end and cell.fill.patternType is None:
                    cell.fill = today_fill

    apply_common_styles(ws)
    ws.freeze_panes = ws.cell(row=4, column=timeline_start_col)
    ws.auto_filter.ref = f"A3:{get_column_letter(timeline_start_col + len(timeline) - 1)}3"

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "$1:$3"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    return ws


def write_raw_sheet(wb: Workbook, issues: List[Issue]) -> Worksheet:
    ws = wb.create_sheet("Issues_Raw")
    headers = [
        "ID",
        "Project",
        "ProjectPath",
        "Tracker",
        "Status",
        "Priority",
        "Subject",
        "AssignedTo",
        "Version",
        "ParentID",
        "StartDate",
        "DueDate",
        "DoneRatio",
        "EstimatedHours",
        "Relations",
        "CustomFields",
    ]
    ws.append(headers)
    for issue in issues:
        cf_text = "; ".join(f"{k}={v}" for k, v in issue.custom_fields.items())
        ws.append([
            issue.issue_id,
            issue.project_name,
            issue.project_path,
            issue.tracker_name,
            issue.status_name,
            issue.priority_name,
            issue.subject,
            issue.assigned_to,
            issue.fixed_version,
            issue.parent_id or "",
            issue.start_date.isoformat() if issue.start_date else "",
            issue.due_date.isoformat() if issue.due_date else "",
            issue.done_ratio,
            issue.estimated_hours if issue.estimated_hours is not None else "",
            relation_text(issue),
            cf_text,
        ])

    widths = {
        1: 8, 2: 20, 3: 28, 4: 14, 5: 14, 6: 12, 7: 50, 8: 18, 9: 16,
        10: 10, 11: 12, 12: 12, 13: 10, 14: 12, 15: 35, 16: 45,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    apply_common_styles(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.sheet_view.showGridLines = False
    return ws


def main() -> int:
    args = parse_args()
    config = read_ini_config(args.config)

    try:
        base_url, api_key = resolve_connection_settings(args, config)
    except ValueError as e:
        print(str(e), file=sys.stderr)
        return 2

    client = RedmineClient(base_url, api_key)
    issues = client.fetch_issues(
        project_id=args.project_id,
        status_id=args.status_id,
        limit=args.limit,
        trackers=args.tracker_id,
        assigned_to_id=args.assigned_to_id,
        version_id=args.fixed_version_id,
        subproject_mode=args.subproject_id,
        fallback_relations=args.fallback_relations,
        sleep_sec=args.sleep_sec,
    )

    if not issues:
        print("No issues found for the given filters.", file=sys.stderr)
        return 1

    projects = client.fetch_projects(limit=args.limit, sleep_sec=args.sleep_sec)
    project_paths = build_project_paths(projects)
    attach_project_paths(issues, project_paths)

    ordered = compute_hierarchy(issues)
    filtered, start, end = detect_date_range(ordered, include_no_date_issues=args.include_no_date_issues)
    start, end = override_date_range(start, end, args.from_date, args.to_date)
    timeline = build_timeline(start, end, args.timeline_mode)

    wb = Workbook()
    write_gantt_sheet(wb, args.sheet_name, filtered, timeline, args.timeline_mode)
    write_raw_sheet(wb, ordered)
    wb.save(args.output)

    print(f"Saved: {args.output}")
    print(f"Issues: {len(ordered)}")
    print(f"Timeline: {start.isoformat()} -> {end.isoformat()} ({args.timeline_mode})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
